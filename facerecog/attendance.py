"""
Concurrent-safe attendance logging to attendance.xlsx.

Two separate processes (in-gate, out-gate) may write at the same time,
so every read-modify-write is wrapped in an flock on a sidecar .lock
file to avoid corrupting the workbook.
"""

import datetime
import fcntl
import os

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ATTENDANCE_PATH = os.path.join(BASE_DIR, "attendance.xlsx")
LOCK_PATH = os.path.join(BASE_DIR, "attendance.xlsx.lock")

HEADERS = ["Employee ID", "Name", "Date", "In Time", "Out Time"]


class _FileLock:
    def __enter__(self):
        self._fh = open(LOCK_PATH, "w")
        fcntl.flock(self._fh, fcntl.LOCK_EX)
        return self

    def __exit__(self, *exc):
        fcntl.flock(self._fh, fcntl.LOCK_UN)
        self._fh.close()


def _open_workbook():
    if os.path.exists(ATTENDANCE_PATH):
        wb = openpyxl.load_workbook(ATTENDANCE_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(HEADERS)
    return wb, ws


def log_event(employee_id, name, mode, when=None):
    """Record an in/out event. mode is 'in' or 'out'.

    IN only fills the first blank In Time for the day (later INs the
    same day are ignored). OUT always overwrites Out Time with the
    latest event, so the sheet reflects the most recent exit.
    Returns True if the sheet was actually modified.
    """
    when = when or datetime.datetime.now()
    date_str = when.strftime("%Y-%m-%d")
    time_str = when.strftime("%H:%M:%S")

    with _FileLock():
        wb, ws = _open_workbook()

        target_row = None
        for row in ws.iter_rows(min_row=2):
            if row[0].value == employee_id and row[2].value == date_str:
                target_row = row
                break

        changed = False

        if target_row is None:
            new_row = [employee_id, name, date_str, None, None]
            if mode == "in":
                new_row[3] = time_str
            else:
                new_row[4] = time_str
            ws.append(new_row)
            changed = True
        else:
            if mode == "in":
                if target_row[3].value in (None, ""):
                    target_row[3].value = time_str
                    changed = True
            else:
                target_row[4].value = time_str
                changed = True

        if changed:
            wb.save(ATTENDANCE_PATH)

        return changed
